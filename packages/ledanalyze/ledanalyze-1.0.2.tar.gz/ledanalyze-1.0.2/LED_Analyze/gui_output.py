"""
LED_Analyze - gui_output
Date: 15 June 2018
By: Ivan Pougatchev
Version: 1.0
"""
from tkinter import *
from tkinter import messagebox, filedialog
from os.path import exists, splitext
import openpyxl as xl
import matplotlib.pyplot as plt
from matplotlib.colors import BoundaryNorm, Normalize, LinearSegmentedColormap
from matplotlib.ticker import MaxNLocator
import matplotlib.patches as patches
import matplotlib.path as path
import datetime

from gui_common import *
from db import *
        
class InputKeyFrame(LabelFrame):
    """
    Frame that will allow the user to select a single input key for
    data preview or generation of a single report.
    """
    def __init__(self, master):
        
        LabelFrame.__init__(self, master)
        
        # Configure Frame
        self.config(text="Data Selection From Database")
        self.rowconfigure(1, minsize=5)
        
        self.input_key_in = ComboboxFrame(
            self,
            "",
            [],
            480)
        self.input_key_in.grid(
            row=0, column=0, sticky=W)
    
    def update(self):
        self.input_key_in.combox.configure(values=self.master.input_keys)
        
        if len(self.master.input_keys) != 0:
            self.input_key_in.combox_input.set(
                self.master.input_keys[0])
        else:
            self.clear()
    
    def clear(self):
        self.input_key_in.combox_input.set("")

class ReportTemplateFrame(LabelFrame):
    """
    Frame to be placed on the main GUI screen that allows the user to
    set up an output report template.
    """
    def __init__(self, master):
        
        LabelFrame.__init__(self, master)
        
        # Configure Frame
        self.config(text="Report Template Setup")
        self.columnconfigure(0, minsize=5)
        self.columnconfigure(3, minsize=5)
        
        # Initialize Attributes
        self.scalar_vals = self.master.scalar_vals
        self.array_vals = self.master.array_vals
        
        # Add Widgets - Report label selection
        self.label_in = []
        for i in range (0, 2):
            for j in range(0, 2):
                self.label_in.append(
                    ComboboxFrame(
                        self,
                        "Report Label " + str(j+1+2*i) + ":",
                        [],
                        220, 75))
                self.label_in[-1].grid(
                    row=1+i, column=1+j, padx=5, pady=2, sticky=W)
        
        self.lum_map_temp = HeatmapTemplate(
            self,
            "Luminance Heatmap")
        self.lum_map_temp.grid(
            row=3, column=1, columnspan=4, pady=2, sticky=W)
        
        self.lum_hist_temp = Hist1DTemplate(
            self,
            "Luminance Histogram")
        self.lum_hist_temp.grid(
            row=4, column=1, columnspan=4, pady=2, sticky=W)
        
        self.uni_map_temp = HeatmapTemplate(
            self,
            "Uniformity Heatmap")
        self.uni_map_temp.grid(
            row=5, column=1, columnspan=4, pady=2, sticky=W)
        
        self.uni_hist_temp = Hist1DTemplate(
            self,
            "Uniformity Histogram")
        self.uni_hist_temp.grid(
            row=6, column=1, columnspan=4, pady=2, sticky=W)
        
        self.ll_temp = HeatmapTemplate(
            self,
            "Light Leakage Heatmap",
            True)
        self.ll_temp.grid(
            row=7, column=1, columnspan=4, pady=2, sticky=W)
        
        self.chrom_hist_temp = Hist2DTemplate(
            self,
            "Chromaticity Histogram")
        self.chrom_hist_temp.grid(
            row=8, column=1, columnspan=4, pady=2, sticky=W)
        
        self.wl_map_temp = HeatmapTemplate(
            self,
            "Dominant Wavelength Heatmap")
        self.wl_map_temp.grid(
            row=9, column=1, columnspan=4, pady=2, sticky=W)
        
        self.wl_hist_temp = Hist1DTemplate(
            self,
            "Dominant Wavelength Histogram")
        self.wl_hist_temp.grid(
            row=10, column=1, columnspan=4, pady=2, sticky=W)
        
        self.ph_hist_temp = Hist1DTemplate(
            self,
            "Purple Hue Histogram")
        self.ph_hist_temp.grid(
            row=11, column=1, columnspan=4, pady=2, sticky=W)
        
        self.elements = [
            self.lum_map_temp,
            self.lum_hist_temp,
            self.uni_map_temp,
            self.uni_hist_temp,
            self.ll_temp,
            self.chrom_hist_temp,
            self.wl_map_temp,
            self.wl_hist_temp,
            self.ph_hist_temp]
    
    def update(self):
        
        # Update drop down selections for all comboboxes in report template
        # menus
        
        for obj in self.label_in:
            obj.combox.configure(values=self.master.scalar_vals)
        
        self.lum_map_temp.arr_name.combox.configure(
            values=self.master.array_vals)
        
        self.lum_hist_temp.arr_name.combox.configure(
            values=self.master.array_vals)
        
        self.uni_map_temp.arr_name.combox.configure(
            values=self.master.array_vals)
        
        self.uni_hist_temp.arr_name.combox.configure(
            values=self.master.array_vals)
        
        self.ll_temp.arr_name.combox.configure(
            values=self.master.array_vals)
        self.ll_temp.ll_name.combox.configure(
            values=self.master.scalar_vals)
        
        self.chrom_hist_temp.xbins_name.combox.configure(
            values=self.master.array_vals)
        self.chrom_hist_temp.ybins_name.combox.configure(
            values=self.master.array_vals)
        self.chrom_hist_temp.count_name.combox.configure(
            values=self.master.array_vals)
        
        self.wl_map_temp.arr_name.combox.configure(
            values=self.master.array_vals)
        
        self.wl_hist_temp.arr_name.combox.configure(
            values=self.master.array_vals)
        
        self.ph_hist_temp.arr_name.combox.configure(
            values=self.master.array_vals)
        
    def clear(self):
        
        # Clear out all selections in report template frame
        for obj in self.label_in:
            obj.combox_input.set("")
        
        self.lum_map_temp.arr_name.combox_input.set("")
        
        self.lum_hist_temp.arr_name.combox_input.set("")
        
        self.uni_map_temp.arr_name.combox_input.set("")
        
        self.uni_hist_temp.arr_name.combox_input.set("")
        
        self.ll_temp.arr_name.combox_input.set("")
        self.ll_temp.ll_name.combox_input.set("")
        
        self.chrom_hist_temp.xbins_name.combox_input.set("")
        self.chrom_hist_temp.ybins_name.combox_input.set("")
        self.chrom_hist_temp.count_name.combox_input.set("")
        
        self.wl_map_temp.arr_name.combox_input.set("")
        
        self.wl_hist_temp.arr_name.combox_input.set("")
        
        self.ph_hist_temp.arr_name.combox_input.set("")

class ReportElementTemplate(LabelFrame):
    """
    Parent class for all report element frames.
    """
    def __init__(self, master, title):
        
        LabelFrame.__init__(self, master)
        
        # Initialize Attributes
        self.input_key = ""
        self.db_path = ""
        self.return_fig = False
        
        # Configure Frame
        self.config(text=title)
        self.rowconfigure(3, minsize=4)
        
        # Look for changes to input_key and db_path
        self.master.master.input_key_frame.input_key_in.combox_input.trace(
            "w", self.update_input_key)
        self.master.master.db_file_frame.path.trace(
            "w", self.update_db)
        
        # Add Widgets
        self.check_var = IntVar(0)
        Checkbutton(
            self,
            variable=self.check_var).grid(
                row=0, column=1, rowspan=3)
        
        Button(
            self,
            width=16,
            text="Preview",
            command=self.preview).grid(
                row=0, column=3, rowspan=3, padx=5, sticky=W)
            
    def generate_label(self):
        
        label_src = [obj.combox_input.get() for obj in self.master.label_in \
                     if obj.combox_input.get() != ""]
        
        # Open database and retrieve labels
        if len(label_src) > 0:
            analysis_db = AnalysisDatabase(self.db_path)
            label = [str(analysis_db.get_val(self.input_key, header)) \
                     for header in label_src]
            analysis_db.close()
            return "_".join(label)
        else:
            return ""
    
    def update_db(self, *args):
        self.db_path = self.master.master.db_file_frame.path.get()
    
    def update_input_key(self, *args):
        self.input_key = \
            self.master.master.input_key_frame.input_key_in.combox_input.get()

class HeatmapTemplate(ReportElementTemplate):
    
    def __init__(self, master, title, include_ll=False):
        
        ReportElementTemplate.__init__(self, master, title)
        self.include_ll = include_ll
        self.title = title
        
        # Configure Frame
        self.columnconfigure(0, minsize=5)
        self.columnconfigure(1, minsize=25)
        
        # Add Widgets
        self.arr_name = ComboboxFrame(
            self,
            "Array Input:",
            [],
            300, 75)
        self.arr_name.grid(
            row=0, column=2, padx=5, pady=2, sticky=W)
        
        if include_ll:
            self.ll_name = ComboboxFrame(
                self,
                "Light Leak:",
                [],
                300, 75)
            self.ll_name.grid(
                row=1, column=2, padx=5, pady=2, sticky=W)
    
    def generate_plot(self):
        
        # Get label
        label = self.generate_label()
        
        # Check all inputs
        if len(self.master.master.input_keys) == 0:
            messagebox.showerror(
                "No Input Data Sets",
                "No input data sets were found. Either there is no database "
                + "selected, or the selected database has no valid data. "
                + "You should really do something about that.")
        
        arr_name = self.arr_name.combox_input.get()
        if self.include_ll:
            ll_name = self.ll_name.combox_input.get()
        else:
            ll_name = None
        
        if arr_name == "" or ll_name == "":
            messagebox.showerror(
                "Missing Input",
                "Make sure all enabled output plots have a selected input.")
        
        # Get data from database
        analysis_db = AnalysisDatabase(self.db_path)
        z = analysis_db.get_val(
            self.input_key,
            arr_name)
        
        if z.shape[1] == 2:
            messagebox.showwarning(
                "Data Looks Weird",
                "The shape of this data makes it look like you may have "
                + "selected a histogram. Are you sure you didn't?")
        
        if self.include_ll:
            leak = analysis_db.get_val(
                self.input_key,
                ll_name)
        
        # Get noise threshold for heatmap by sampling outer edge of data
        noise_sample = np.concatenate((
            z[0,:],
            z[-1,:],
            z[:,0][1:-1],
            z[:,-1][1:-1]))
        noise_threshold = np.average(noise_sample) + 6 * np.std(noise_sample)
        
        # Setup heatmap
        x = np.arange(np.shape(z)[1])
        y = np.arange(np.shape(z)[0])
        z = z[:-1, :-1]
        
        if self.title == "Dominant Wavelength Heatmap":
            cmap = LinearSegmentedColormap('Spectrum', spec_dict_wl)
            max_val = 830
            min_val = 380
            cmap.set_under('white')
            cmap.set_over('gray')
        else:
            cmap = LinearSegmentedColormap('Rainbow', spec_dict_rb)
            max_val = np.max(z)
            min_val = noise_threshold
        plt.register_cmap(cmap=cmap)
        
        fig, (ax0) = plt.subplots(nrows=1)
        im = ax0.pcolormesh(
            x, y, z, cmap=cmap, 
            norm=Normalize(vmin=min_val, vmax=max_val))
        cbar = fig.colorbar(im, ax=ax0)
        if self.title == "Luminance Heatmap" or \
                self.title == "Light Leakage Heatmap":
            cbar.set_label('Luminance [cd/m^2]', rotation=270, labelpad=15)
        elif self.title == "Uniformity Heatmap":
            cbar.set_label('Uniformity [cd/m^2/mm]', rotation=270, labelpad=15)
        elif self.title == "Dominant Wavelength Heatmap":
            cbar.set_label('Wavelength [nm]', rotation=270, labelpad=15)
        ax0.set_title(self.title + " - " + label)
        
        if self.include_ll:
            plt.text(int(z.shape[1]/10), int(z.shape[0]/10), 
                     "Leakage Ratio: " + str(leak))
        
        fig.tight_layout()
        
        if self.return_fig:
            return fig
    
    def preview(self):
        self.generate_plot()
        plt.show()

class Hist1DTemplate(ReportElementTemplate):
    
    def __init__(self, master, title):
        
        ReportElementTemplate.__init__(self, master, title)
        self.title = title
        
        # Configure Frame
        self.columnconfigure(0, minsize=5)
        self.columnconfigure(1, minsize=25)
        
        # Add Widgets
        self.arr_name = ComboboxFrame(
            self,
            "Hist Input:",
            [],
            300, 75)
        self.arr_name.grid(
            row=0, column=2, padx=5, pady=2, sticky=W)
    
    def generate_plot(self):
        
        # Get label
        label = self.generate_label()
        
        # Check all inputs
        if len(self.master.master.input_keys) == 0:
            messagebox.showerror(
                "No Input Data Sets",
                "No input data sets were found. Either there is no database "
                + "selected, or the selected database has no valid data. "
                + "You should really do something about that.")
        
        arr_name = self.arr_name.combox_input.get()
        
        if arr_name == "":
            messagebox.showerror(
                "Missing Input",
                "Make sure all enabled output plots have a selected input.")
            
        analysis_db = AnalysisDatabase(self.db_path)
        arr = analysis_db.get_val(self.input_key, arr_name)
        
        if arr.shape[1] != 2:
            messagebox.showerror(
                "Data Looks Weird",
                "This is most certainly not a histogram. Try again!")
        
        # Unzip data from database check for logarithmic scale and
        # fix last bin
        bins, n = arr[:,0], arr[:,1]
        dx1 = bins[-1] - bins[-2]
        dx2 = bins[-2] - bins[-3]
        if dx1 / dx2 > 1.01:
            make_log = True
            bins = np.append(bins, bins[-1] + dx1 * (dx1/dx2))
        else:
            make_log = False
            bins = np.append(bins, bins[-1] + dx1)
            
        # Setup histogram
        fig, (ax0) = plt.subplots(nrows=1)
        left = np.array(bins[:-1])
        right = np.array(bins[1:])
        bottom = np.zeros(len(left))
        top = bottom + n
        
        XY = np.array([[left,left,right,right], [bottom,top,top,bottom]]).T
        barpath = path.Path.make_compound_path_from_polys(XY)
        patch = patches.PathPatch(
            barpath, facecolor=(0.49, 0.15, 0.80), edgecolor='white', alpha=0.8)
        ax0.add_patch(patch)
        
        if make_log:
            ax0.semilogx()
        ax0.set_xlim(left[0], right[-1])
        ax0.set_ylim(bottom.min(), top.max())
        if self.title == "Luminance Histogram":
            ax0.set_xlabel('Luminance [cd/m^2]')
        elif self.title == "Uniformity Histogram":
            ax0.set_xlabel('Uniformity [cd/m^2/mm]')
        elif self.title == "Dominant Wavelength Histogram":
            ax0.set_xlabel('Wavelength [nm]')
        ax0.set_title(self.title + " - " + label)
        fig.tight_layout()
        
        if self.return_fig:
            return fig
    
    def preview(self):
        self.generate_plot()
        plt.show()

class Hist2DTemplate(ReportElementTemplate):
    
    def __init__(self, master, title):
        
        ReportElementTemplate.__init__(self, master, title)
        self.title = title
        
        # Configure Frame
        self.columnconfigure(0, minsize=5)
        self.columnconfigure(1, minsize=25)
        
        # Add Widgets
        self.xbins_name = ComboboxFrame(
            self,
            "X Bins Input:",
            [],
            300, 75)
        self.xbins_name.grid(
            row=0, column=2, padx=5, pady=2, sticky=W)
        
        self.ybins_name = ComboboxFrame(
            self,
            "Y Bins Input:",
            [],
            300, 75)
        self.ybins_name.grid(
            row=1, column=2, padx=5, pady=2, sticky=W)
        
        self.count_name = ComboboxFrame(
            self,
            "Count Input:",
            [],
            300, 75)
        self.count_name.grid(
            row=2, column=2, padx=5, pady=2, sticky=W)
    
    def generate_plot(self):
    
        # Get label
        label = self.generate_label()
        
        # Check all inputs
        if len(self.master.master.input_keys) == 0:
            messagebox.showerror(
                "No Input Data Sets",
                "No input data sets were found. Either there is no database "
                + "selected, or the selected database has no valid data. "
                + "You should really do something about that.")
        
        xbins_name = self.xbins_name.combox_input.get()
        ybins_name = self.ybins_name.combox_input.get()
        count_name = self.count_name.combox_input.get()
        
        if "" in (xbins_name, ybins_name, count_name):
            messagebox.showerror(
                "Missing Input",
                "Make sure all enabled output plots have a selected input.")
            
        analysis_db = AnalysisDatabase(self.db_path)
        xbins = analysis_db.get_val(self.input_key, xbins_name)
        ybins = analysis_db.get_val(self.input_key, ybins_name)
        count = analysis_db.get_val(self.input_key, count_name)
        
        if len(xbins.shape) != 1 or len(ybins.shape) != 1:
            messagebox.showerror(
                "Data Looks Weird",
                "The X bins or Y bins of the chromaticity histogram are not"
                + " 1-D arrays. Sad!")
        
        if xbins.shape[0]-1 != count.shape[0] or \
                ybins.shape[0]-1 != count.shape[1]:
            messagebox.showerror(
                "Data Looks Weird",
                "The shape of the chromaticity histogram does not match the "
                + "shape of the bin arrays.")
        
        # Setup histogram plot
        cmap = LinearSegmentedColormap('Purply', spec_dict_purp)
        fig, (ax0) = plt.subplots(nrows=1)
        x, y = np.meshgrid(xbins, ybins)
        im = ax0.pcolormesh(x, y, count, cmap=cmap)
        cbar = fig.colorbar(im, ax=ax0)
        ax0.set_xlabel('x-chromaticity coordinate')
        ax0.set_ylabel('y-chromaticity coordinate', rotation=90, labelpad=10)
        ax0.set_title(self.title + " - " + label)
        
        if self.return_fig:
            return fig
    
    def preview(self):
        self.generate_plot()
        plt.show()

class OutputDataFrame(Frame):
    """
    Frame to be placed on the main GUI screen containing all options to
    output calculated data.
    """
    def __init__(self, master):
        
        Frame.__init__(self, master)
        
        # Add Widgets
        Button(
            self,
            text="Output to Spreadsheet",
            width=20,
            command=self.output_to_ss).grid(
                row=0, column=0, padx=5, pady=5, sticky=N+S)
        
        Button(
            self,
            text="Output Single Report",
            width=20,
            command=self.output_single_set).grid(
                    row=0, column=1, padx=5, pady=5, sticky=N+S)
        
        Button(
            self,
            text="Output All Reports",
            width=20,
            command=self.output_all_sets).grid(
                row=0, column=2, padx=5, pady=5, sticky=N+S)
        
    def output_to_ss(self):
        
        if exists(self.master.db_path.get()):
            
            # Open analysis database
            analysis_db = AnalysisDatabase(
                self.master.db_path.get())
            
            # Create output file
            output_path = filedialog.asksaveasfilename(
                title="Create data summary spreadsheet",
                initialdir="C:\\",
                filetypes=[("Excel File", "*.xlsx")])
            
            name, tail = splitext(output_path)
            output_path = name + ".xlsx"
             
            # Prepare new excel file for data insertion
            wb = xl.Workbook()
            ws = wb.worksheets[0]
             
            # Get column headers and row keys from database table
            headers = analysis_db.get_data_headers()
            
            headers = [header for header in headers \
                       if analysis_db.get_col_type(header) == "Scalar"]
            keys = analysis_db.get_keys()
            
            for i, header in enumerate(headers):
                ws.cell(column=i+1, row=1).value = header
                for j, key in enumerate(keys):
                    ws.cell(column=i+1, row=j+2).value = \
                        analysis_db.get_val(key, header)
             
            wb.save(output_path)
            
        else:
            messagebox.showerror(
                "Database Not Loaded",
                "Please load a valid database file before attempting to "
                + "output.")

    def output_single_set(self):
        
        output_path = filedialog.askdirectory(
            title="Choose an Ouput Directory",
            initialdir="C:\\") 
        
        self.generate_set(output_path)
    
    def output_all_sets(self):
        
        output_path = filedialog.askdirectory(
            title="Choose an Ouput Directory",
            initialdir="C:\\") 
        
        for input_key in self.master.input_keys:
            self.master.input_key_frame.input_key_in.combox_input.set(
                input_key)
            self.generate_set(output_path)
    
    def generate_set(self, path):
        for item in self.master.report_temp_frame.elements:
            if item.check_var.get() == 1:
                output_name = item.title.replace(" ", "_") + "_" \
                    + item.generate_label() + "_" \
                    + "{:%Y-%m-%d-%H-%M-%S}".format(datetime.datetime.now()) \
                    + ".png"
                item.return_fig = True 
                fig = item.generate_plot()
                item.return_fig = False
                fig.savefig(path + "\\" + output_name)
                plt.close('all')

### Custom colormap for heatmap data
spec_dict_rb = {'red': ((0, 1, 1),
                        (0.071, 0, 0),
                        (0.286, 0, 0),
                        (0.357, 0, 0),
                        (0.571, 1, 1),
                        (0.929, 1, 1),
                        (1, 0.5, 0.5)),
                
                'green': ((0, 1, 1),
                          (0.071, 0, 0),
                          (0.286, 1, 1),
                          (0.357, 1, 1),
                          (0.571, 1, 1),
                          (0.929, 0, 0),
                          (1, 0, 0)),
                
                'blue': ((0, 1, 1),
                         (0.071, 1, 1),
                         (0.286, 1, 1),
                         (0.357, 0, 0),
                         (0.571, 0, 0),
                         (0.929, 0, 0),
                         (1, 0, 0))}

### Custom colormap for wavelength data
spec_dict_wl = {'red': ((0, 0.3804, 0.3804),
                        (0.0426, 0.3804, 0.3804),
                        (0.0638, 0.4745, 0.4745),
                        (0.0851, 0.5137, 0.5137),
                        (0.0915, 0.5137, 0.5137),
                        (0.1064, 0.4941, 0.4941),
                        (0.1277, 0.4157, 0.4157),
                        (0.1702, 0, 0),
                        (0.3191, 0, 0),
                        (0.4043, 0.6392, 0.6392),
                        (0.4681, 1, 1),
                        (0.7191, 1, 1),
                        (0.8936, 0.3804, 0.3804),
                        (1, 0.3804, 0.3804)),
                
                'green': ((0, 0, 0),
                          (0.0426, 0, 0),
                          (0.1702, 0, 0),
                          (0.2234, 0.5725, 0.5725),
                          (0.2766, 1, 1),
                          (0.4681, 1, 1),
                          (0.4894, 0.8745, 0.8745),
                          (0.5532, 0.4667, 0.4667),
                          (0.617, 0, 0),
                          (1, 0, 0)),
                
                'blue': ((0, 0.3804, 0.3804),
                         (0.0426, 0.3804, 0.3804),
                         (0.0638, 0.5529, 0.5529),
                         (0.0851, 0.7098, 0.7098),
                         (0.0915, 0.7529, 0.7529),
                         (0.1064, 0.8588, 0.8588),
                         (0.1277, 1, 1),
                         (0.2766, 1, 1),
                         (0.3191, 0, 0),
                         (1, 0, 0))}

### Custom colormap for 2D histogram(0.49, 0.15, 0.80)
spec_dict_purp = {'red': ((0, 1, 1),
                          (1, 0.49, 0.49)),
                          
                  'green': ((0, 1, 1),
                            (1, 0.15, 0.15)),
                  
                  'blue': ((0, 1, 1),
                           (1, 0.80, 0.80))}